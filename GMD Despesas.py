import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Despesas", page_icon="📊", layout="wide")

st.markdown("""<style>
.main-header{background:linear-gradient(90deg,#1F3864,#2E75B6);color:white;
  padding:20px 30px;border-radius:10px;margin-bottom:20px}
.main-header h1{color:white;margin:0;font-size:1.8rem}
.main-header p{color:#cce0f5;margin:4px 0 0 0;font-size:.9rem}
.metric-card{background:white;border:1px solid #e0e0e0;border-radius:8px;
  padding:16px;text-align:center;box-shadow:0 2px 4px rgba(0,0,0,.05)}
.metric-label{font-size:.75rem;color:#666;font-weight:600;text-transform:uppercase}
.metric-value{font-size:1.35rem;font-weight:700;margin-top:4px}
.metric-ok{color:#375623} .metric-attn{color:#9C0006} .metric-blue{color:#1F3864}
.sec-hdr{background:#1F3864;color:white;padding:8px 16px;border-radius:6px;
  margin:20px 0 10px 0;font-weight:700;font-size:.95rem}
.stFileUploader section button{background-color:#1F3864 !important;color:white !important;border:none !important;border-radius:8px !important;font-weight:600 !important}
.stFileUploader section button:hover{background-color:#2E75B6 !important}
.stDownloadButton button{background-color:#1F3864 !important;color:white !important;border:none !important;border-radius:8px !important;font-weight:600 !important}
.stDownloadButton button:hover{background-color:#2E75B6 !important}
div[data-baseweb=select]>div{background-color:#1F3864 !important;color:white !important;border-radius:8px !important;font-weight:600 !important}
div[data-baseweb=select] svg{fill:white !important}
div[data-baseweb=popover] li{color:#1F3864 !important}
div[data-baseweb=popover] li:hover{background-color:#D6E4F0 !important}
</style>""", unsafe_allow_html=True)

MESES = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']

def brl(v):
    if pd.isna(v): return "-"
    s = "R$ {:,.2f}".format(abs(v)).replace(",","X").replace(".",",").replace("X",".")
    return "({})".format(s) if v < 0 else s

def pct(v):
    return "-" if (pd.isna(v) or v == 0) else "{:.1f}%".format(v * 100)

def detectar_ano(orc_bytes):
    tmp = pd.read_csv(io.BytesIO(orc_bytes), encoding='latin1', sep=None, engine='python', nrows=0)
    for col in tmp.columns:
        if 'Planejado' in col and '/' in col:
            return col.split('/')[1]
    return '26'

@st.cache_data
def load(orc_bytes, pac_bytes):
    df_orc = pd.read_csv(io.BytesIO(orc_bytes), encoding='latin1', sep=None, engine='python')
    df_pac = pd.read_excel(io.BytesIO(pac_bytes))
    df_orc['Estr. da conta'] = df_orc['Estr. da conta'].astype(str).str.replace('="','').str.replace('"','').str.strip()
    df_pac['Estrutura da conta'] = df_pac['Estrutura da conta'].astype(str).str.strip()
    mc = [c for c in df_orc.columns if 'Planejado' in c or 'Realizado' in c]
    for col in mc:
        df_orc[col] = pd.to_numeric(df_orc[col].astype(str).str.replace('.','',regex=False).str.replace(',','.',regex=False), errors='coerce').fillna(0)
    merged = df_orc.merge(df_pac[['Estrutura da conta','Nome do pacote','Gestor do pacote']], left_on='Estr. da conta', right_on='Estrutura da conta', how='inner')
    if 'Descricao da conta' not in merged.columns and 'Descrição da conta' in merged.columns:
        merged = merged.rename(columns={'Descrição da conta':'Descricao da conta'})
    return merged, mc

def resumo(merged, mc, mes, ano):
    pc = ["Planejado{}/{}".format(m, ano) for m in MESES]
    idx = MESES.index(mes)
    mr = MESES[:idx+1]
    grp = merged.groupby(['Nome do pacote','Estr. da conta','Descricao da conta','Gestor do pacote'], as_index=False)[mc].sum()
    rows = []
    for pac in sorted(grp['Nome do pacote'].unique()):
        s = grp[grp['Nome do pacote']==pac]
        g = s['Gestor do pacote'].iloc[0]
        gn = g.split('@')[0].replace('.',' ').title() if '@' in g else g
        pm = "Planejado{}/{}".format(mes, ano)
        rm2 = "Realizado{}/{}".format(mes, ano)
        om = s[pm].sum(); rmv = s[rm2].sum()
        oa = s[pc].sum().sum()
        ra = s[["Realizado{}/{}".format(m, ano) for m in mr]].sum().sum()
        fut = s[["Planejado{}/{}".format(m, ano) for m in MESES[idx+1:]]].sum().sum()
        tend = ra + fut; tvo = tend - oa
        rows.append({'Pacote':pac,'Gestor':gn,
                     'Orcado Mes':om,'Realizado Mes':rmv,
                     '% Exec Mes':rmv/om if om!=0 else None,
                     'Saldo Mes':rmv-om,'Orcado Anual':oa,
                     'Real Acumulado':ra,'Tendencia Anual':tend,
                     '% Exec Anual':tend/oa if oa!=0 else None,
                     'Tend Vs Orcamento':tvo,
                     'Status':'OK' if tvo>=0 else 'Atencao'})
    return pd.DataFrame(rows), grp

def detalhe(grp, mc, mes, pac, ano):
    pc = ["Planejado{}/{}".format(m, ano) for m in MESES]
    idx = MESES.index(mes)
    mr = MESES[:idx+1]
    sub = grp[grp['Nome do pacote']==pac]
    rows = []
    for _, r in sub.iterrows():
        pm = "Planejado{}/{}".format(mes, ano)
        rm2 = "Realizado{}/{}".format(mes, ano)
        om = r[pm]; rmv = r[rm2]
        oa = sum(r[p] for p in pc)
        ra = sum(r["Realizado{}/{}".format(m, ano)] for m in mr)
        fut = sum(r["Planejado{}/{}".format(m, ano)] for m in MESES[idx+1:])
        tend = ra + fut; tvo = tend - oa
        rows.append({'Estrutura':r['Estr. da conta'],
                     'Descricao':r['Descricao da conta'],
                     'Orcado Mes':om,'Realizado Mes':rmv,
                     '% Exec':rmv/om if om!=0 else None,
                     'Saldo Mes':rmv-om,'Orcado Anual':oa,
                     'Real Acum':ra,'Tendencia':tend,
                     'Tend Vs Orcamento':tvo,
                     'Status':'OK' if tvo>=0 else 'Atencao'})
    return pd.DataFrame(rows)

def gerar_excel(grp, mc, mes, ano):
    FN='Arial'; CP='1F3864'; CH='2E75B6'; L1='F2F7FC'; L2='FFFFFF'
    pc=["Planejado{}/{}".format(m,ano) for m in MESES]
    idx=MESES.index(mes); mr=MESES[:idx+1]
    pacotes=sorted(grp['Nome do pacote'].unique())
    def fl(h): return PatternFill('solid',start_color=h,fgColor=h)
    def bd():
        s=Side(style='thin',color='CCCCCC'); return Border(left=s,right=s,top=s,bottom=s)
    def bw():
        s=Side(style='thin',color='FFFFFF'); return Border(left=s,right=s,top=s,bottom=s)
    NF='#,##0.00;(#,##0.00);"-"'; PF='0.0%;(0.0%);"-"'
    bm=Border(left=Side(style='thin',color='FFFFFF'),right=Side(style='thin',color='FFFFFF'),
              top=Side(style='thin',color='FFFFFF'),bottom=Side(style='thin',color='FFFFFF'))

    wb=Workbook()

    # ── RESUMO ────────────────────────────────────────────────────────────────
    ws=wb.active; ws.title='RESUMO'
    ws.sheet_view.showGridLines=False; ws.freeze_panes='A5'
    ws.merge_cells('A1:L2')
    ws['A1']='ACOMPANHAMENTO DE DESPESAS {} - RESUMO POR PACOTE'.format('20'+ano)
    ws['A1'].font=Font(name=FN,bold=True,size=13,color='FFFFFF'); ws['A1'].fill=fl(CP)
    ws['A1'].alignment=Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height=22; ws.row_dimensions[2].height=6
    ws.merge_cells('A3:L3')
    ws['A3']='Mes: {}/{} | Tendencia = Realizado acumulado + Orcado restante'.format(mes,ano)
    ws['A3'].font=Font(name=FN,italic=True,size=9,color='FFFFFF'); ws['A3'].fill=fl('2E75B6')
    ws['A3'].alignment=Alignment(horizontal='center',vertical='center'); ws.row_dimensions[3].height=16
    hdrs=['Pacote','Gestor','Orcado {}/{}'.format(mes,ano),'Realizado {}/{}'.format(mes,ano),
          '% Exec Mes','Saldo Mes','Orcado Anual','Realizado Acum.',
          'Tendencia Anual','% Exec Anual','Tend Vs Orcamento','Status']
    wds=[20,30,16,16,11,16,16,16,16,12,20,10]
    for j,(h,w) in enumerate(zip(hdrs,wds),1):
        c=ws.cell(row=4,column=j,value=h); c.font=Font(name=FN,bold=True,color='FFFFFF',size=10)
        c.fill=fl(CH); c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        c.border=bd(); ws.column_dimensions[get_column_letter(j)].width=w
    ws.row_dimensions[4].height=30
    row=5
    for i,pac in enumerate(pacotes):
        s=grp[grp['Nome do pacote']==pac]; g=s['Gestor do pacote'].iloc[0]
        gn=g.split('@')[0].replace('.',' ').title() if '@' in g else g
        pm="Planejado{}/{}".format(mes,ano); rm2="Realizado{}/{}".format(mes,ano)
        om=s[pm].sum(); rmv=s[rm2].sum()
        oa=s[pc].sum().sum()
        ra=s[["Realizado{}/{}".format(m,ano) for m in mr]].sum().sum()
        fut=s[["Planejado{}/{}".format(m,ano) for m in MESES[idx+1:]]].sum().sum()
        tend=ra+fut; tvo=tend-oa; bg=L1 if i%2==0 else L2
        vals=[pac,gn,om,rmv,
              '=IF(C{r}=0,"-",D{r}/C{r})'.format(r=row),
              '=D{r}-C{r}'.format(r=row),
              oa,ra,tend,
              '=IF(G{r}=0,"-",I{r}/G{r})'.format(r=row),
              tvo,'OK' if tvo>=0 else 'Atencao']
        for j,v in enumerate(vals,1):
            c=ws.cell(row=row,column=j,value=v); c.border=bd()
            if j in(3,4,6,7,8,9,11):
                c.number_format=NF; c.alignment=Alignment(horizontal='right',vertical='center')
                c.font=Font(name=FN,size=9); c.fill=fl(bg)
            elif j in(5,10):
                c.number_format=PF; c.alignment=Alignment(horizontal='center',vertical='center')
                c.font=Font(name=FN,size=9); c.fill=fl(bg)
            elif j==12:
                c.alignment=Alignment(horizontal='center',vertical='center')
                if tvo>=0: c.fill=fl('C6EFCE'); c.font=Font(name=FN,bold=True,size=9,color='375623')
                else: c.fill=fl('FFC7CE'); c.font=Font(name=FN,bold=True,size=9,color='9C0006')
            else:
                c.alignment=Alignment(horizontal='left',vertical='center')
                c.font=Font(name=FN,bold=(j==1),size=9); c.fill=fl(bg)
        ws.row_dimensions[row].height=18; row+=1
    ws.merge_cells('A{}:B{}'.format(row,row)); ws['A{}'.format(row)]='TOTAL GERAL'
    ws['A{}'.format(row)].font=Font(name=FN,bold=True,size=9,color='FFFFFF')
    ws['A{}'.format(row)].fill=fl(CP)
    ws['A{}'.format(row)].alignment=Alignment(horizontal='center',vertical='center')
    for j in range(1,13):
        c=ws.cell(row=row,column=j); c.fill=fl(CP); c.border=bd(); cl=get_column_letter(j)
        if j in(3,4,6,7,8,9,11):
            c.value='=SUM({}5:{}{})'.format(cl,cl,row-1)
            c.font=Font(name=FN,bold=True,size=9,color='FFFFFF')
            c.number_format=NF; c.alignment=Alignment(horizontal='right',vertical='center')
        elif j==5:
            c.value='=IF(C{r}=0,"-",D{r}/C{r})'.format(r=row); c.number_format=PF
            c.font=Font(name=FN,bold=True,size=9,color='FFFFFF')
            c.alignment=Alignment(horizontal='center',vertical='center')
        elif j==10:
            c.value='=IF(G{r}=0,"-",I{r}/G{r})'.format(r=row); c.number_format=PF
            c.font=Font(name=FN,bold=True,size=9,color='FFFFFF')
            c.alignment=Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[row].height=20

    # ── ABAS POR PACOTE ───────────────────────────────────────────────────────
    for pac in pacotes:
        sub=grp[grp['Nome do pacote']==pac].copy().reset_index(drop=True)
        ws2=wb.create_sheet(title=pac[:31]); ws2.sheet_view.showGridLines=False; ws2.freeze_panes='A5'
        nc=11; ws2.merge_cells('A1:{}2'.format(get_column_letter(nc)))
        ws2['A1']='PACOTE: '+pac; ws2['A1'].font=Font(name=FN,bold=True,size=12,color='FFFFFF')
        ws2['A1'].fill=fl(CP); ws2['A1'].alignment=Alignment(horizontal='center',vertical='center')
        ws2.row_dimensions[1].height=22; ws2.row_dimensions[2].height=6
        g2=sub['Gestor do pacote'].iloc[0] if len(sub) else ''
        ws2.merge_cells('A3:{}3'.format(get_column_letter(nc)))
        ws2['A3']='Gestor: {}  |  Mes: {}/{}'.format(g2,mes,ano)
        ws2['A3'].font=Font(name=FN,italic=True,size=9,color='FFFFFF'); ws2['A3'].fill=fl('2E75B6')
        ws2['A3'].alignment=Alignment(horizontal='center',vertical='center'); ws2.row_dimensions[3].height=15
        dh=['Estrutura','Descricao','Orcado {}'.format(mes),'Realizado {}'.format(mes),
            '% Exec','Saldo Mes','Orcado Anual','Real Acum','Tendencia','Tend Vs Orcamento','Status']
        dw=[18,40,15,15,10,15,15,15,15,20,12]
        for j,(h,w) in enumerate(zip(dh,dw),1):
            c=ws2.cell(row=4,column=j,value=h); c.font=Font(name=FN,bold=True,color='FFFFFF',size=10)
            c.fill=fl(CH); c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
            c.border=bd(); ws2.column_dimensions[get_column_letter(j)].width=w
        ws2.row_dimensions[4].height=28; rd=5
        for i,r in sub.iterrows():
            pm="Planejado{}/{}".format(mes,ano); rm2="Realizado{}/{}".format(mes,ano)
            om2=r[pm]; rmv2=r[rm2]
            oa2=sum(r[p] for p in pc)
            ra2=sum(r["Realizado{}/{}".format(m,ano)] for m in mr)
            fut2=sum(r["Planejado{}/{}".format(m,ano)] for m in MESES[idx+1:])
            tend2=ra2+fut2; tvo2=tend2-oa2; bg=L1 if i%2==0 else L2
            rv=[r['Estr. da conta'],r['Descricao da conta'],om2,rmv2,
                rmv2/om2 if om2!=0 else 0,rmv2-om2,oa2,ra2,tend2,tvo2,
                'OK' if tvo2>=0 else 'Atencao']
            for j,v in enumerate(rv,1):
                c=ws2.cell(row=rd,column=j,value=v); c.fill=fl(bg); c.border=bd()
                if j in(3,4,6,7,8,9,10):
                    c.number_format=NF; c.alignment=Alignment(horizontal='right',vertical='center')
                    c.font=Font(name=FN,bold=(j==10),size=9)
                elif j==5:
                    c.number_format=PF; c.alignment=Alignment(horizontal='center',vertical='center')
                    c.font=Font(name=FN,size=9)
                elif j==11:
                    c.alignment=Alignment(horizontal='center',vertical='center')
                    if tvo2>=0: c.fill=fl('C6EFCE'); c.font=Font(name=FN,bold=True,size=9,color='375623')
                    else: c.fill=fl('FFC7CE'); c.font=Font(name=FN,bold=True,size=9,color='9C0006')
                else:
                    c.alignment=Alignment(horizontal='left',vertical='center')
                    c.font=Font(name=FN,size=9)
            ws2.row_dimensions[rd].height=16; rd+=1
        ws2.merge_cells('A{}:B{}'.format(rd,rd)); ws2['A{}'.format(rd)]='TOTAL'
        ws2['A{}'.format(rd)].font=Font(name=FN,bold=True,size=9,color='FFFFFF')
        ws2['A{}'.format(rd)].fill=fl(CP)
        ws2['A{}'.format(rd)].alignment=Alignment(horizontal='center',vertical='center')
        for j in range(1,nc+1):
            c=ws2.cell(row=rd,column=j); c.fill=fl(CP); c.border=bd(); cl=get_column_letter(j)
            if j in(3,4,6,7,8,9,10):
                c.value='=SUM({}5:{}{})'.format(cl,cl,rd-1)
                c.font=Font(name=FN,bold=True,size=9,color='FFFFFF')
                c.number_format=NF; c.alignment=Alignment(horizontal='right',vertical='center')
            elif j==5:
                c.value='=IF(C{r}=0,"-",D{r}/C{r})'.format(r=rd); c.number_format=PF
                c.font=Font(name=FN,bold=True,size=9,color='FFFFFF')
                c.alignment=Alignment(horizontal='center',vertical='center')
        ws2.row_dimensions[rd].height=20

    # ── ABA CONSOLIDADO ───────────────────────────────────────────────────────
    wc=wb.create_sheet(title='CONSOLIDADO')
    wc.sheet_view.showGridLines=False; wc.freeze_panes='E6'
    n_cols_c = 4 + len(MESES)*3
    wc.merge_cells('A1:{}'.format(get_column_letter(n_cols_c))+'2')
    wc['A1']='CONSOLIDADO GERAL - ORCADO x REALIZADO x VARIACAO - 20'+ano
    wc['A1'].font=Font(name=FN,bold=True,size=13,color='FFFFFF'); wc['A1'].fill=fl(CP)
    wc['A1'].alignment=Alignment(horizontal='center',vertical='center')
    wc.row_dimensions[1].height=22; wc.row_dimensions[2].height=6
    wc.merge_cells('A3:'+get_column_letter(n_cols_c)+'3')
    wc['A3']='AH R$ = Realizado - Orcado  |  Positivo = Economia  |  Negativo = Estouro'
    wc['A3'].font=Font(name=FN,italic=True,size=9,color='FFFFFF'); wc['A3'].fill=fl('2E75B6')
    wc['A3'].alignment=Alignment(horizontal='center',vertical='center'); wc.row_dimensions[3].height=14
    # Cabecalho linha 4 - meses agrupados
    for ci in range(1,5):
        wc.cell(row=4,column=ci).fill=fl(CP); wc.cell(row=4,column=ci).border=bd()
    col_s=5
    for m in MESES:
        cl1=get_column_letter(col_s); cl3=get_column_letter(col_s+2)
        wc.merge_cells('{}4:{}4'.format(cl1,cl3))
        ch=wc.cell(row=4,column=col_s,value='{}/{}'.format(m,ano))
        ch.font=Font(name=FN,bold=True,color='FFFFFF',size=10)
        ch.fill=fl('2E75B6'); ch.alignment=Alignment(horizontal='center',vertical='center')
        for ci_m in range(col_s, col_s+3):
            is_last=(ci_m==col_s+2)
            wc.cell(row=4,column=ci_m).fill=fl('2E75B6')
            wc.cell(row=4,column=ci_m).border=Border(
                left=Side(style='medium',color='FFFFFF') if ci_m==col_s else Side(style='thin',color='FFFFFF'),
                right=Side(style='medium',color='FFFFFF') if is_last else Side(style='thin',color='FFFFFF'),
                top=Side(style='medium',color='FFFFFF'),
                bottom=Side(style='thin',color='FFFFFF'))
        col_s+=3
    wc.row_dimensions[4].height=20
    # Cabecalho linha 5 - subcolumnas
    for ci,lbl in enumerate(['Pacote','Gestor','Estrutura','Descricao'],1):
        c=wc.cell(row=5,column=ci,value=lbl)
        c.font=Font(name=FN,bold=True,color='FFFFFF',size=9)
        c.fill=fl(CP); c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=bd()
    wc.column_dimensions['A'].width=18; wc.column_dimensions['B'].width=22
    wc.column_dimensions['C'].width=14; wc.column_dimensions['D'].width=38
    col_s=5
    for m in MESES:
        for ci2,lbl in enumerate(['Orcado','Realizado','AH R$'],col_s):
            c=wc.cell(row=5,column=ci2,value=lbl)
            c.font=Font(name=FN,bold=True,color='FFFFFF',size=8)
            borda_sub=Border(left=Side(style='thin',color='CCCCCC'),right=Side(style='medium',color='FFFFFF') if lbl=='AH R$' else Side(style='thin',color='CCCCCC'),top=Side(style='thin',color='CCCCCC'),bottom=Side(style='thin',color='CCCCCC'))
            c.fill=fl(CH); c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=borda_sub
            wc.column_dimensions[get_column_letter(ci2)].width=13
        col_s+=3
    wc.row_dimensions[5].height=22
    # Dados
    row_c=6
    for pac in pacotes:
        sub=grp[grp['Nome do pacote']==pac].reset_index(drop=True)
        g=sub['Gestor do pacote'].iloc[0] if len(sub) else ''
        gn=g.split('@')[0].replace('.',' ').title() if '@' in g else g
        for i_r,r in sub.iterrows():
            bg=L1 if row_c%2==0 else L2
            for ci,v in enumerate([pac,gn,r['Estr. da conta'],r['Descricao da conta']],1):
                c=wc.cell(row=row_c,column=ci,value=v)
                c.font=Font(name=FN,size=8,bold=(ci==1))
                c.fill=fl(bg); c.border=bd()
                c.alignment=Alignment(horizontal='left' if ci>2 else 'center',vertical='center')
            col_d=5
            for m in MESES:
                orc_v=r['Planejado{}/{}'.format(m,ano)]
                real_v=r['Realizado{}/{}'.format(m,ano)]
                ah_v=real_v-orc_v
                for ci2,v in enumerate([orc_v,real_v,ah_v],col_d):
                    is_ah=(ci2==col_d+2)
                    c=wc.cell(row=row_c,column=ci2,value=v)
                    c.number_format=NF; c.fill=fl(bg)
                    c.border=Border(left=Side(style="thin",color="CCCCCC"),right=Side(style="medium",color="FFFFFF") if is_ah else Side(style="thin",color="CCCCCC"),top=Side(style="thin",color="CCCCCC"),bottom=Side(style="thin",color="CCCCCC"))
                    c.alignment=Alignment(horizontal="right",vertical="center")
                    if is_ah:
                        if v<0: c.font=Font(name=FN,size=8,color='9C0006',bold=True)
                        elif v>0: c.font=Font(name=FN,size=8,color='375623',bold=True)
                        else: c.font=Font(name=FN,size=8)
                    else: c.font=Font(name=FN,size=8)
                col_d+=3
            wc.row_dimensions[row_c].height=14; row_c+=1
    # Total Geral Consolidado
    wc.merge_cells('A{}:D{}'.format(row_c,row_c))
    wc['A{}'.format(row_c)]='TOTAL GERAL'
    wc['A{}'.format(row_c)].font=Font(name=FN,bold=True,size=9,color='FFFFFF')
    wc['A{}'.format(row_c)].fill=fl(CP)
    wc['A{}'.format(row_c)].alignment=Alignment(horizontal='center',vertical='center')
    col_d=5
    for m in MESES:
        for ci2 in range(col_d,col_d+3):
            cl=get_column_letter(ci2)
            c=wc.cell(row=row_c,column=ci2)
            c.value='=SUM({}6:{}{})'.format(cl,cl,row_c-1)
            c.font=Font(name=FN,bold=True,size=9,color='FFFFFF')
            c.fill=fl(CP); c.number_format=NF; c.border=bd()
            c.alignment=Alignment(horizontal='right',vertical='center')
        col_d+=3
    for ci in range(1,4):
        wc.cell(row=row_c,column=ci).fill=fl(CP)
        wc.cell(row=row_c,column=ci).border=bd()
    wc.row_dimensions[row_c].height=20

    # ── MENU ─────────────────────────────────────────────────────────────────
    wm=wb.create_sheet(title='MENU',index=0); wm.sheet_view.showGridLines=False
    wm.sheet_view.showRowColHeaders=False
    wm.column_dimensions['A'].width=4; wm.column_dimensions['B'].width=36; wm.column_dimensions['C'].width=28
    wm.merge_cells('B2:C3')
    wm['B2']='ACOMPANHAMENTO DE DESPESAS {}'.format('20'+ano)
    wm['B2'].font=Font(name=FN,bold=True,size=16,color='FFFFFF'); wm['B2'].fill=fl(CP)
    wm['B2'].alignment=Alignment(horizontal='center',vertical='center')
    wm.row_dimensions[2].height=28; wm.row_dimensions[3].height=28
    wm.merge_cells('B4:C4'); wm['B4']='Mes: {}/{}'.format(mes,ano)
    wm['B4'].font=Font(name=FN,italic=True,size=9,color='FFFFFF'); wm['B4'].fill=fl('2E75B6')
    wm['B4'].alignment=Alignment(horizontal='center',vertical='center'); wm.row_dimensions[4].height=16
    wm.row_dimensions[5].height=10
    wm.merge_cells('B6:C6'); wm['B6']='  VISOES GERAIS'
    wm['B6'].font=Font(name=FN,bold=True,size=9,color='FFFFFF'); wm['B6'].fill=fl('2E75B6')
    wm['B6'].alignment=Alignment(horizontal='left',vertical='center'); wm.row_dimensions[6].height=18
    # Resumo
    c7=wm.cell(row=7,column=2,value='Resumo por Pacote')
    c7.font=Font(name=FN,bold=True,size=10,color='FFFFFF'); c7.fill=fl('2E75B6')
    c7.alignment=Alignment(horizontal='left',vertical='center',indent=1); c7.border=bm; c7.hyperlink='#RESUMO!A1'
    d7=wm.cell(row=7,column=3,value='Visao consolidada de todos os pacotes')
    d7.font=Font(name=FN,size=9,color='1F3864'); d7.fill=fl('D6E4F0')
    d7.alignment=Alignment(horizontal='left',vertical='center',indent=1); d7.border=bm
    wm.row_dimensions[7].height=24
    # Consolidado
    c8=wm.cell(row=8,column=2,value='Consolidado Geral')
    c8.font=Font(name=FN,bold=True,size=10,color='FFFFFF'); c8.fill=fl('1A7A4A')
    c8.alignment=Alignment(horizontal='left',vertical='center',indent=1); c8.border=bm; c8.hyperlink='#CONSOLIDADO!A1'
    d8=wm.cell(row=8,column=3,value='Orcado x Realizado x AH por mes')
    d8.font=Font(name=FN,size=9,color='1F3864'); d8.fill=fl('D6E4F0')
    d8.alignment=Alignment(horizontal='left',vertical='center',indent=1); d8.border=bm
    wm.row_dimensions[8].height=24
    c9=wm.cell(row=9,column=2,value='Desvios Significativos')
    c9.font=Font(name=FN,bold=True,size=10,color='FFFFFF'); c9.fill=fl('9C0006')
    c9.alignment=Alignment(horizontal='left',vertical='center',indent=1); c9.border=bm; c9.hyperlink='#DESVIOS!A1'
    d9=wm.cell(row=9,column=3,value='Contas com desvio acima de R$ 5.000')
    d9.font=Font(name=FN,size=9,color='1F3864'); d9.fill=fl('D6E4F0')
    d9.alignment=Alignment(horizontal='left',vertical='center',indent=1); d9.border=bm
    wm.row_dimensions[9].height=24
    wm.row_dimensions[10].height=10
    wm.merge_cells('B11:C11'); wm['B11']='  PACOTES'
    wm['B11'].font=Font(name=FN,bold=True,size=9,color='FFFFFF'); wm['B11'].fill=fl(CP)
    wm['B11'].alignment=Alignment(horizontal='left',vertical='center'); wm.row_dimensions[11].height=18
    gm=grp.groupby('Nome do pacote')['Gestor do pacote'].first().to_dict()
    CB=['1A5276','1F618D','2471A3','2980B9','5499C8','7FB3D3','2E86C1','1B4F72','154360','1A5276','117A65']
    rm3=12
    for ki,pac in enumerate(pacotes):
        gr2=gm.get(pac,''); gn2=gr2.split('@')[0].replace('.',' ').title() if '@' in gr2 else gr2
        cb2=wm.cell(row=rm3,column=2,value='PACOTE: '+pac)
        cb2.font=Font(name=FN,bold=True,size=10,color='FFFFFF'); cb2.fill=fl(CB[ki%len(CB)])
        cb2.alignment=Alignment(horizontal='left',vertical='center',indent=1); cb2.border=bm
        cb2.hyperlink='#'+pac[:31]+'!A1'
        cd2=wm.cell(row=rm3,column=3,value='Gestor: '+gn2)
        cd2.font=Font(name=FN,size=9,color='1F3864'); cd2.fill=fl('EBF5FB')
        cd2.alignment=Alignment(horizontal='left',vertical='center',indent=1); cd2.border=bm
        wm.row_dimensions[rm3].height=24; rm3+=1
    wm.row_dimensions[rm3].height=10; rm3+=1
    wm.merge_cells('B{}:C{}'.format(rm3,rm3))
    wm['B{}'.format(rm3)]='Clique em qualquer item para navegar diretamente a aba correspondente.'
    wm['B{}'.format(rm3)].font=Font(name=FN,italic=True,size=8,color='7F7F7F')
    wm['B{}'.format(rm3)].alignment=Alignment(horizontal='center',vertical='center')
    wm.row_dimensions[rm3].height=16

    def ab(wt):
        nc=wt.max_column+1; bc=wt.cell(row=1,column=nc,value='Home Menu')
        bc.hyperlink='#MENU!A1'; bc.font=Font(name=FN,bold=True,size=9,color='FFFFFF')
        bc.fill=fl(CP); bc.alignment=Alignment(horizontal='center',vertical='center')
        bc.border=bm; wt.column_dimensions[get_column_letter(nc)].width=12

    # ── ABA DESVIOS ───────────────────────────────────────────────────────────
    LIMITE = 5000
    wd = wb.create_sheet(title='DESVIOS')
    wd.sheet_view.showGridLines = False
    wd.freeze_panes = 'A6'

    wd.merge_cells('A1:H2')
    wd['A1'] = 'DESVIOS SIGNIFICATIVOS - AH R$ ACIMA DE R$ 5.000'
    wd['A1'].font = Font(name=FN, bold=True, size=13, color='FFFFFF')
    wd['A1'].fill = fl('9C0006')
    wd['A1'].alignment = Alignment(horizontal='center', vertical='center')
    wd.row_dimensions[1].height = 22; wd.row_dimensions[2].height = 6

    wd.merge_cells('A3:H3')
    wd['A3'] = 'Desvios do mes {}/{} | Positivo = Economia | Negativo = Estouro'.format(mes, ano)
    wd['A3'].font = Font(name=FN, italic=True, size=9, color='FFFFFF')
    wd['A3'].fill = fl('CC0000')
    wd['A3'].alignment = Alignment(horizontal='center', vertical='center')
    wd.row_dimensions[3].height = 14; wd.row_dimensions[4].height = 6

    hdrs_d = ['Pacote','Gestor','Estrutura','Descricao','Orcado Mes','Realizado Mes','AH R$','Justificativa']
    wds_d  = [18, 22, 14, 40, 16, 16, 16, 45]
    for j,(h,w) in enumerate(zip(hdrs_d, wds_d), 1):
        c = wd.cell(row=5, column=j, value=h)
        c.font = Font(name=FN, bold=True, color='FFFFFF', size=9)
        c.fill = fl('9C0006') if h != 'Justificativa' else fl('1F3864')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = bd()
        wd.column_dimensions[get_column_letter(j)].width = w
    wd.row_dimensions[5].height = 22

    pm_d = 'Planejado{}/{}'.format(mes, ano)
    rm_d = 'Realizado{}/{}'.format(mes, ano)
    row_d = 6
    for pac in pacotes:
        sub = grp[grp['Nome do pacote'] == pac].reset_index(drop=True)
        g = sub['Gestor do pacote'].iloc[0] if len(sub) else ''
        gn = g.split('@')[0].replace('.', ' ').title() if '@' in g else g
        for _, r in sub.iterrows():
            om_d = r[pm_d]; rm2_d = r[rm_d]; ah_d = rm2_d - om_d
            if abs(ah_d) < LIMITE: continue
            bg     = 'F0FFF0' if ah_d > 0 else 'FFF0F0'
            ah_cor = '375623' if ah_d > 0 else '9C0006'
            vals_d = [pac, gn, r['Estr. da conta'], r['Descricao da conta'], om_d, rm2_d, ah_d, '']
            for j, v in enumerate(vals_d, 1):
                c = wd.cell(row=row_d, column=j, value=v)
                c.fill = fl(bg); c.border = bd()
                c.alignment = Alignment(vertical='center', horizontal='right' if j in (5,6,7) else 'left', wrap_text=(j==8))
                if j in (5,6,7):
                    c.number_format = NF
                    if j == 7: c.font = Font(name=FN, size=9, bold=True, color=ah_cor)
                    else:      c.font = Font(name=FN, size=9)
                elif j == 8:
                    c.fill = fl('FFFDE7')
                    c.font = Font(name=FN, size=9, italic=True, color='999999')
                    c.value = 'Digite aqui...'
                else:
                    c.font = Font(name=FN, size=9, bold=(j==1))
            wd.row_dimensions[row_d].height = 18
            row_d += 1

    if row_d > 6:
        wd.merge_cells('A{}:D{}'.format(row_d, row_d))
        wd['A{}'.format(row_d)] = 'TOTAL ({} desvios)'.format(row_d - 6)
        wd['A{}'.format(row_d)].font = Font(name=FN, bold=True, size=9, color='FFFFFF')
        wd['A{}'.format(row_d)].fill = fl('9C0006')
        wd['A{}'.format(row_d)].alignment = Alignment(horizontal='center', vertical='center')
        for j in range(1, 9):
            c = wd.cell(row=row_d, column=j); c.fill = fl('9C0006'); c.border = bd()
            cl = get_column_letter(j)
            if j in (5,6,7):
                c.value = '=SUM({}6:{}{})'.format(cl, cl, row_d-1)
                c.font = Font(name=FN, bold=True, size=9, color='FFFFFF')
                c.number_format = NF
                c.alignment = Alignment(horizontal='right', vertical='center')
        wd.row_dimensions[row_d].height = 20
    else:
        wd.merge_cells('A6:H6')
        wd['A6'] = 'Nenhum desvio acima de R$ 5.000 encontrado no mes {}/{}'.format(mes, ano)
        wd['A6'].font = Font(name=FN, italic=True, size=10, color='666666')
        wd['A6'].alignment = Alignment(horizontal='center', vertical='center')
        wd.row_dimensions[6].height = 30

    ab(wd)
    ab(ws); ab(wc)
    for pac in pacotes: ab(wb[pac[:31]])
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ─── INTERFACE ────────────────────────────────────────────────────────────────
st.markdown("""<div class="main-header">
  <h1>📊 Acompanhamento de Despesas</h1>
  <p>Orcado x Realizado por Pacote - Tendencia Anual e Saldo Disponivel</p>
</div>""", unsafe_allow_html=True)

with st.sidebar:
    st.markdown("## Configuracoes")
    st.markdown("---")
    st.markdown("**Arquivos**")
    orc_file = st.file_uploader("Orcamento (CSV)", type=['csv'], key="orc_upload")
    pac_file = st.file_uploader("Tabela de Pacotes (XLSX)", type=['xlsx'], key="pac_upload")
    st.markdown("---")
    st.markdown("**Mes de referencia**")
    if orc_file:
        ano_ref = detectar_ano(orc_file.read())
        orc_file.seek(0)
    else:
        ano_ref = '26'
    idx_mes = st.selectbox("Mes de referencia", options=[1,2,3,4,5,6,7,8,9,10,11,12], index=1, label_visibility="collapsed")
    mes_sel = MESES[idx_mes - 1]
    st.markdown("---")
    st.markdown("""<small style='color:#888'>Como usar:<br>
    1. Faca upload dos dois arquivos<br>2. Selecione o mes<br>
    3. Explore o resumo e detalhes<br>4. Baixe o Excel gerado</small>""", unsafe_allow_html=True)

if orc_file and pac_file:
    orc_bytes = orc_file.read()
    pac_bytes = pac_file.read()
    with st.spinner("Processando dados..."):
        try:
            merged, mc = load(orc_bytes, pac_bytes)
            df_res, grp = resumo(merged, mc, mes_sel, ano_ref)
        except Exception as e:
            st.error("Erro ao processar: {}".format(e)); st.stop()

    tom=df_res['Orcado Mes'].sum(); trm=df_res['Realizado Mes'].sum()
    ttvo=df_res['Tend Vs Orcamento'].sum()
    n_ok=(df_res['Status']=='OK').sum(); n_at=(df_res['Status']=='Atencao').sum()
    pem=trm/tom if tom!=0 else 0

    k1,k2,k3,k4,k5=st.columns(5)
    with k1: st.markdown('<div class="metric-card"><div class="metric-label">Orcado {}/{}</div><div class="metric-value metric-blue">{}</div></div>'.format(mes_sel,ano_ref,brl(tom)),unsafe_allow_html=True)
    with k2: st.markdown('<div class="metric-card"><div class="metric-label">Realizado {}/{}</div><div class="metric-value metric-blue">{}</div></div>'.format(mes_sel,ano_ref,brl(trm)),unsafe_allow_html=True)
    with k3:
        cor="metric-attn" if pem>1.05 else "metric-ok"
        st.markdown('<div class="metric-card"><div class="metric-label">% Execucao Mes</div><div class="metric-value {}">{}</div></div>'.format(cor,pct(pem)),unsafe_allow_html=True)
    with k4:
        ct="metric-ok" if ttvo>=0 else "metric-attn"
        st.markdown('<div class="metric-card"><div class="metric-label">Tend. Vs Orcamento</div><div class="metric-value {}">{}</div></div>'.format(ct,brl(ttvo)),unsafe_allow_html=True)
    with k5: st.markdown('<div class="metric-card"><div class="metric-label">Pacotes</div><div class="metric-value"><span class="metric-ok">OK {}</span> <span class="metric-attn">Atencao {}</span></div></div>'.format(n_ok,n_at),unsafe_allow_html=True)

    st.markdown("<br>",unsafe_allow_html=True)
    st.markdown('<div class="sec-hdr">📋 Resumo por Pacote</div>',unsafe_allow_html=True)

    cn=['Orcado Mes','Realizado Mes','Saldo Mes','Orcado Anual','Real Acumulado','Tendencia Anual','Tend Vs Orcamento']
    cp2=['% Exec Mes','% Exec Anual']

    def sty(df):
        s=pd.DataFrame('',index=df.index,columns=df.columns)
        for i,r in df.iterrows():
            if r['Status']=='OK':
                s.at[i,'Status']='background-color:#C6EFCE;color:#375623;font-weight:bold;text-align:center'
                s.at[i,'Tend Vs Orcamento']='color:#375623;font-weight:bold'
            else:
                s.at[i,'Status']='background-color:#FFC7CE;color:#9C0006;font-weight:bold;text-align:center'
                s.at[i,'Tend Vs Orcamento']='color:#9C0006;font-weight:bold'
        return s

    fd={c:brl for c in cn}; fd.update({c:pct for c in cp2})
    st.dataframe(
        df_res.style.format(fd).apply(sty,axis=None)
          .set_properties(**{'font-size':'13px'})
          .set_table_styles([
              {'selector':'thead th','props':[('background-color','#2E75B6'),('color','white'),('font-weight','bold'),('text-align','center'),('padding','8px 12px')]},
              {'selector':'tbody tr:nth-child(even)','props':[('background-color','#F2F7FC')]},
              {'selector':'tbody tr:hover','props':[('background-color','#ddeeff')]},
          ]),
        use_container_width=True,hide_index=True
    )

    st.markdown('<div class="sec-hdr">🔍 Detalhe por Pacote</div>',unsafe_allow_html=True)
    pac_sel=st.selectbox("Pacote",sorted(df_res['Pacote'].unique()),label_visibility="collapsed")
    df_d=detalhe(grp,mc,mes_sel,pac_sel,ano_ref)

    def sty2(df):
        s=pd.DataFrame('',index=df.index,columns=df.columns)
        for i,r in df.iterrows():
            if r['Status']=='OK':
                s.at[i,'Status']='background-color:#C6EFCE;color:#375623;font-weight:bold;text-align:center'
                s.at[i,'Tend Vs Orcamento']='color:#375623;font-weight:bold'
            else:
                s.at[i,'Status']='background-color:#FFC7CE;color:#9C0006;font-weight:bold;text-align:center'
                s.at[i,'Tend Vs Orcamento']='color:#9C0006;font-weight:bold'
        return s

    cn2=['Orcado Mes','Realizado Mes','Saldo Mes','Orcado Anual','Real Acum','Tendencia','Tend Vs Orcamento']
    fd2={c:brl for c in cn2}; fd2['% Exec']=pct
    st.dataframe(
        df_d.style.format(fd2).apply(sty2,axis=None)
          .set_properties(**{'font-size':'12px'})
          .set_table_styles([
              {'selector':'thead th','props':[('background-color','#2E75B6'),('color','white'),('font-weight','bold'),('text-align','center'),('padding','6px 10px')]},
              {'selector':'tbody tr:nth-child(even)','props':[('background-color','#F2F7FC')]},
          ]),
        use_container_width=True,hide_index=True
    )

    st.markdown("---")
    col1,col2=st.columns([1,3])
    with col1:
        xls=gerar_excel(grp,mc,mes_sel,ano_ref)
        st.download_button("Baixar Relatorio Excel",data=xls,
            file_name="Acompanhamento_Despesas_{}_{}.xlsx".format(mes_sel,'20'+ano_ref),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with col2:
        st.caption("Excel com MENU + RESUMO + CONSOLIDADO + {} abas de pacote.".format(len(df_res)))

else:
    st.markdown("""<div style='text-align:center;padding:60px 20px;color:#666'>
        <div style='font-size:4rem'>📂</div>
        <h3 style='color:#1F3864'>Faca o upload dos arquivos para comecar</h3>
        <p>Use o painel lateral para carregar <b>ORCAMENTO (CSV)</b> e <b>PACOTES (XLSX)</b></p>
        <br>
        <div style='display:flex;justify-content:center;gap:40px;margin-top:20px'>
            <div style='background:#f0f7ff;border:2px dashed #2E75B6;border-radius:10px;padding:20px 30px'>
                <div style='font-size:2rem'>📄</div><b>ORCAMENTO.csv</b><br>
                <small>Planejado e Realizado por mes</small></div>
            <div style='background:#f0f7ff;border:2px dashed #2E75B6;border-radius:10px;padding:20px 30px'>
                <div style='font-size:2rem'>📦</div><b>PACOTES.xlsx</b><br>
                <small>Cadastro de pacotes e responsaveis</small></div>
        </div></div>""", unsafe_allow_html=True)
